import os
import pandas as pd


def test_allocations_csv_matches_portfolio_sheet(tmp_path):
    # run fixture pipeline to create allocations.csv and excel
    cfg = os.path.abspath("example-config-fixtures.yaml")
    import subprocess

    subprocess.run(["python", "scripts/ci_fixture_run.py", cfg], check=True)
    alloc_path = os.path.abspath("output/allocations.csv")
    excel_path = os.path.abspath("output/fixture_macro_ranking.xlsx")
    assert os.path.exists(alloc_path)
    assert os.path.exists(excel_path)
    # read allocations csv
    alloc = pd.read_csv(alloc_path)
    assert "country" in alloc.columns and "weight" in alloc.columns
    # read excel Portfolio sheet if present
    from openpyxl import load_workbook

    wb = load_workbook(excel_path, read_only=True)
    sheets = wb.sheetnames
    # Portfolio sheet may be present
    if "Portfolio" in sheets:
        ws = wb["Portfolio"]
        rows = list(ws.values)
        headers = rows[0]
        # find country and weight columns
        idx_country = headers.index("country") if "country" in headers else None
        idx_weight = headers.index("weight") if "weight" in headers else None
        assert idx_country is not None and idx_weight is not None
        portfolio_df = pd.DataFrame(rows[1:], columns=headers)
        # compare sums
        csv_sum = float(alloc["weight"].sum())
        port_sum = float(pd.to_numeric(portfolio_df["weight"]).sum())
        assert abs(csv_sum - port_sum) < 1e-6
    else:
        # if no Portfolio sheet, at least allocations CSV should exist and be non-empty
        assert len(alloc) > 0
