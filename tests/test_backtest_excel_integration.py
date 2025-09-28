import pandas as pd
import tempfile
import os


def test_export_excel_includes_backtest_and_portfolio():
    from src.io.excel import export_to_excel

    # create small ranking, indicators, raw and backtest/portfolio frames
    ranking = pd.DataFrame({"country": ["USA", "DEU"], "score": [0.9, 0.1]}).set_index("country")
    indicators = pd.DataFrame({"country": ["USA", "DEU"], "indicator": ["gdp", "gdp"], "value": [1.0, 0.5]})
    raw = pd.DataFrame({"country": ["USA", "DEU"], "date": [pd.Timestamp("2020-01-01"), pd.Timestamp("2020-01-01")], "value": [1.0, 0.5]})

    backtest_df = pd.DataFrame({"date": pd.date_range("2020-01-01", periods=3), "nav": [1.0, 1.01, 1.02]})
    portfolio_df = pd.DataFrame({"country": ["USA", "DEU"], "weight": [0.6, 0.4]})

    cfg = {"excel": {"number_format": "#.##0,00"}, "backtest": {"results": backtest_df}, "portfolio": {"allocations": portfolio_df}}

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    tmp.close()
    path = tmp.name
    try:
        out = export_to_excel(path, ranking, indicators, raw, cfg)
        assert os.path.exists(out)
        # basic check: open with openpyxl and confirm sheets
        from openpyxl import load_workbook

        wb = load_workbook(out, read_only=True)
        names = wb.sheetnames
        assert "Backtest" in names or any("Backtest" in n for n in names)
        assert "Portfolio" in names or any("Portfolio" in n for n in names)
    finally:
        try:
            os.remove(path)
        except Exception:
            pass
