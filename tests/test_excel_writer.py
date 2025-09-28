import os
import pandas as pd
from src.io.excel import export_to_excel


def test_export_writes_portfolio_and_backtest(tmp_path):
    # create small ranking, indicators, raw and portfolio/backtest frames
    ranking = pd.DataFrame({"country": ["DEU", "FRA"], "score": [0.8, 0.6]})
    indicators = pd.DataFrame(
        {
            "iso3": ["DEU", "FRA"],
            "indicator": ["gdp", "gdp"],
            "value": [1, 2],
            "date": ["2020-01-01", "2020-01-01"],
        }
    )
    raw = pd.DataFrame(
        {
            "iso3": ["DEU", "FRA"],
            "value": [1.1, 2.2],
            "date": ["2020-01-01", "2020-01-01"],
        }
    )

    portfolio = pd.DataFrame(
        {"country": ["DEU", "FRA"], "weight": [0.6, 0.4], "prev_weight": [0.5, 0.45]}
    )
    backtest = pd.DataFrame(
        {"date": ["2020-01-01", "2020-04-01"], "nav": [100.0, 102.0]}
    )

    cfg = {
        "excel": {},
        "portfolio": {"allocations": portfolio},
        "backtest": {"results": backtest},
    }
    out = tmp_path / "out.xlsx"
    path = export_to_excel(str(out), ranking, indicators, raw, cfg)
    assert os.path.exists(path)
    # open with openpyxl and check sheets exist
    from openpyxl import load_workbook

    wb = load_workbook(path)
    assert "Ranking_Übersicht" in wb.sheetnames
    assert "Portfolio" in wb.sheetnames
    assert "Backtest" in wb.sheetnames
    # check est_cost/delta columns present in Portfolio sheet
    ws = wb["Portfolio"]
    headers = [c.value for c in ws[1]]
    assert "delta" in headers or "est_cost" in headers


def test_export_handles_iso3_mapping(tmp_path):
    ranking = pd.DataFrame({"country": ["DEU", "FRA"], "score": [0.9, 0.1]})
    indicators = pd.DataFrame(
        {
            "iso3": ["DEU", "FRA"],
            "indicator": ["gdp", "gdp"],
            "value": [10, 20],
            "date": ["2020-01-01", "2020-01-01"],
        }
    )
    raw = pd.DataFrame(
        {
            "iso3": ["DEU", "USA"],
            "value": [1.1, 3.3],
            "date": ["2020-01-01", "2020-01-01"],
        }
    )
    cfg = {"excel": {"number_format": "#,##0.00", "date_format": "DD.MM.YYYY"}}
    out = tmp_path / "out2.xlsx"
    path = export_to_excel(str(out), ranking, indicators, raw, cfg)
    from openpyxl import load_workbook

    wb = load_workbook(path)
    # check that sheet exists and header includes score
    ws = wb["Ranking_Übersicht"]
    headers = [c.value for c in ws[1]]
    assert "score" in headers
    # indicators sheet should include 'value' column
    ws2 = wb["Einzelindikatoren"]
    headers2 = [c.value for c in ws2[1]]
    assert "value" in headers2


def test_export_reads_mapping_path_and_writes_ticker(tmp_path):
    # create mapping CSV
    mapping_dir = tmp_path / "data"
    mapping_dir.mkdir()
    mapping_path = mapping_dir / "countries_iso3_map.csv"
    mapping_path.write_text(
        "iso3,ticker,isin,exchange,currency\nDEU,DEUTICK,DEU123,EX1,EUR\nFRA,FRATICK,FRA123,EX2,EUR\n"
    )

    ranking = pd.DataFrame({"country": ["DEU", "FRA"], "score": [0.8, 0.6]})
    indicators = pd.DataFrame(
        {
            "iso3": ["DEU", "FRA"],
            "indicator": ["gdp", "gdp"],
            "value": [1, 2],
            "date": ["2020-01-01", "2020-01-01"],
        }
    )
    raw = pd.DataFrame(
        {
            "iso3": ["DEU", "FRA"],
            "value": [1.1, 2.2],
            "date": ["2020-01-01", "2020-01-01"],
        }
    )

    portfolio = pd.DataFrame(
        {"country": ["DEU", "FRA"], "weight": [0.6, 0.4], "prev_weight": [0.5, 0.45]}
    )
    cfg = {
        "excel": {},
        "portfolio": {"allocations": portfolio, "mapping_path": str(mapping_path)},
        "backtest": None,
    }
    out = tmp_path / "out_map.xlsx"
    path = export_to_excel(str(out), ranking, indicators, raw, cfg)
    from openpyxl import load_workbook

    wb = load_workbook(path)
    assert "Portfolio" in wb.sheetnames
    ws = wb["Portfolio"]
    headers = [c.value for c in ws[1]]
    # mapping should have added a ticker column
    assert "ticker" in headers


def test_export_permission_fallback(tmp_path, monkeypatch):
    # ensure the save permission fallback path is exercised
    ranking = pd.DataFrame({"country": ["DEU"], "score": [0.5]})
    indicators = pd.DataFrame(
        {"iso3": ["DEU"], "indicator": ["gdp"], "value": [1], "date": ["2020-01-01"]}
    )
    raw = pd.DataFrame({"iso3": ["DEU"], "value": [1.1], "date": ["2020-01-01"]})
    cfg = {"excel": {}, "portfolio": None}
    out = tmp_path / "out_locked.xlsx"

    # monkeypatch Workbook.save to raise on first call then succeed
    import openpyxl

    calls = {"n": 0}

    orig_save = openpyxl.workbook.workbook.Workbook.save

    def fake_save(self, path):
        calls["n"] += 1
        if calls["n"] == 1:
            raise PermissionError("file locked")
        # on subsequent calls, delegate to original implementation
        return orig_save(self, path)

    monkeypatch.setattr(openpyxl.workbook.workbook.Workbook, "save", fake_save)

    path = export_to_excel(str(out), ranking, indicators, raw, cfg)
    assert os.path.exists(path)


def test_export_writes_harmonize_report(tmp_path):
    ranking = pd.DataFrame({"country": ["DEU"], "score": [0.5]})
    indicators = pd.DataFrame(
        {"iso3": ["DEU"], "indicator": ["gdp"], "value": [1], "date": ["2020-01-01"]}
    )
    raw = pd.DataFrame({"iso3": ["DEU"], "value": [1.1], "date": ["2020-01-01"]})
    hrep = pd.DataFrame({"indicator": ["gdp"], "rule": ["mean"]})
    cfg = {"excel": {}, "portfolio": None, "harmonize_report": hrep}
    out = tmp_path / "out_hrep.xlsx"
    path = export_to_excel(str(out), ranking, indicators, raw, cfg)
    from openpyxl import load_workbook

    wb = load_workbook(path)
    assert "Harmonize_Report" in wb.sheetnames
