import os
import pandas as pd
from openpyxl import load_workbook

from src.io.excel import export_to_excel


def test_excel_portfolio_mapping(tmp_path, monkeypatch):
    # create data/countries_iso3_map.csv in repo root
    repo_root = os.getcwd()
    data_dir = os.path.join(repo_root, "data")
    os.makedirs(data_dir, exist_ok=True)
    csv_path = os.path.join(data_dir, "countries_iso3_map.csv")
    with open(csv_path, "w", encoding="utf-8") as fh:
        fh.write("iso3,ticker,isin,exchange,currency\n")
        fh.write("USA,SPY,US78462F1030,NYSE,USD\n")
        fh.write("DEU,EWG,US4642871849,NYSE,USD\n")

    ranked = pd.DataFrame({"score": [1.0, 0.5]}, index=["USA", "DEU"]) 
    indicators_df = pd.DataFrame({"country": ["USA","DEU"], "indicator": ["i1","i2"], "date": ["2020-01-01","2020-01-01"], "value": [1,2]})
    raw_df = indicators_df.copy()
    cfg = {"excel": {"path": str(tmp_path / "test.xlsx")}, "portfolio": {"allocations": pd.DataFrame({"country": ["USA","DEU"], "weight": [0.6, 0.4], "prev_weight": [0.5, 0.5]})}}
    out = export_to_excel(str(tmp_path / "test.xlsx"), ranked, indicators_df, raw_df, cfg)
    # open and check Portfolio sheet exists and contains ticker/isin
    wb = load_workbook(out)
    assert "Portfolio" in wb.sheetnames
    ws = wb["Portfolio"]
    headers = [c.value for c in ws[1]]
    assert "ticker" in headers
    assert "isin" in headers
    # delta and est_cost should be present because prev_weight was provided
    assert "delta" in headers
    assert "est_cost" in headers
