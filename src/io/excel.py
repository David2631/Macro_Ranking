import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# numbers format constants are optional; we don't directly reference `numbers` so omit to satisfy linter
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.worksheet.table import Table, TableStyleInfo
import os
from typing import Dict, Optional, Any
import csv
import pycountry


def export_to_excel(
    path: str,
    ranking_df: pd.DataFrame,
    indicators_df: pd.DataFrame,
    raw_df: pd.DataFrame,
    config: Dict[str, Any],
) -> str:
    os.makedirs(os.path.dirname(path), exist_ok=True)
    wb = Workbook()
    # load ISO3 -> country name mapping (try data/countries_iso3.csv first)

    country_map: Dict[str, str] = {}
    try:
        csv_path = os.path.join(os.getcwd(), "data", "countries_iso3.csv")
        if os.path.exists(csv_path):
            with open(csv_path, newline="", encoding="utf-8") as csv_fh:
                reader = csv.reader(csv_fh)
                for row in reader:
                    if not row:
                        continue
                    # accept both with/without header: first column ISO3, second column name
                    code = row[0].strip()
                    name = row[1].strip() if len(row) > 1 else ""
                    if code:
                        country_map[code.upper()] = name or country_map.get(
                            code.upper(), ""
                        )
    except Exception:
        country_map = {}

    def pycountry_lookup(code: str) -> Optional[str]:
        try:
            if not code or pd.isna(code):
                return None
            c = pycountry.countries.get(alpha_3=str(code))
            if c:
                return c.name
        except Exception:
            return None
        return None

    def map_iso3_column_in_df(df: pd.DataFrame) -> pd.DataFrame:
        # operate on a copy
        df = df.copy()
        # prefer explicit column names
        target_cols = [
            c for c in df.columns if str(c).lower() in ("country", "iso3", "code")
        ]
        if not target_cols:
            # fallback: check first column for 3-letter ISO3 codes
            first = df.columns[0]
            sample = df[first].dropna().astype(str)
            if not sample.empty and (sample.str.len() == 3).mean() > 0.5:
                target_cols = [first]
        for col in target_cols:

            def conv(v):
                try:
                    if pd.isna(v):
                        return v
                    s = str(v).upper()
                    if s in country_map and country_map[s]:
                        return country_map[s]
                    pc = pycountry_lookup(s)
                    return pc or v
                except Exception:
                    return v

            df[col] = df[col].apply(conv)
        return df

    # Ranking sheet (map ISO3 -> country name where possible)
    ws = wb.active
    ws.title = "Ranking_Übersicht"
    try:
        r_df = map_iso3_column_in_df(ranking_df.reset_index())
    except Exception:
        r_df = ranking_df.reset_index()
    for r in dataframe_to_rows(r_df, index=False, header=True):
        ws.append(r)
    # Freeze header
    ws.freeze_panes = "A2"
    # Create table
    tab = Table(
        displayName="RankingTable",
        ref=f"A1:{ws.cell(row=ws.max_row, column=ws.max_column).coordinate}",
    )
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
    )
    tab.tableStyleInfo = style
    ws.add_table(tab)
    # Apply number format to score column if present
    number_format = config.get("excel", {}).get("number_format", "#.##0,00")
    date_format = config.get("excel", {}).get("date_format", "DD.MM.YYYY")
    # find score column
    headers = [c.value for c in ws[1]]
    if "score" in headers:
        idx = headers.index("score") + 1
        for cell in ws.iter_cols(min_col=idx, max_col=idx, min_row=2):
            for c in cell:
                c.number_format = number_format
    # also format CI columns if present
    for colname in ("score_ci_low", "score_ci_high", "rank_stability"):
        if colname in headers:
            idx = headers.index(colname) + 1
            for cell in ws.iter_cols(min_col=idx, max_col=idx, min_row=2):
                for c in cell:
                    c.number_format = number_format
    # Conditional formatting for score
    if "score" in headers:
        idx = headers.index("score") + 1
        first = ws.cell(row=2, column=idx).coordinate
        last = ws.cell(row=ws.max_row, column=idx).coordinate
        ws.conditional_formatting.add(
            f"{first}:{last}",
            ColorScaleRule(
                start_type="min",
                start_color="F8696B",
                mid_type="percentile",
                mid_value=50,
                mid_color="FFEB84",
                end_type="max",
                end_color="63BE7B",
            ),
        )
    # Einzelindikatoren
    ws2 = wb.create_sheet("Einzelindikatoren")
    try:
        ind_df = map_iso3_column_in_df(indicators_df.reset_index())
    except Exception:
        ind_df = indicators_df.reset_index()
    for r in dataframe_to_rows(ind_df, index=False, header=True):
        ws2.append(r)
    # format value/date columns if present
    headers2 = [c.value for c in ws2[1]]
    if "value" in headers2:
        idxv = headers2.index("value") + 1
        for cell in ws2.iter_cols(min_col=idxv, max_col=idxv, min_row=2):
            for c in cell:
                c.number_format = number_format
    if "date" in headers2:
        idxd = headers2.index("date") + 1
        for cell in ws2.iter_cols(min_col=idxd, max_col=idxd, min_row=2):
            for c in cell:
                c.number_format = date_format
    ws2.freeze_panes = "A2"
    tab2 = Table(
        displayName="IndicatorTable",
        ref=f"A1:{ws2.cell(row=ws2.max_row, column=ws2.max_column).coordinate}",
    )
    tab2.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    ws2.add_table(tab2)
    # Raw data
    ws3 = wb.create_sheet("Rohdaten")
    try:
        raw_mapped = map_iso3_column_in_df(raw_df.reset_index())
    except Exception:
        raw_mapped = raw_df.reset_index()
    for r in dataframe_to_rows(raw_mapped, index=False, header=True):
        ws3.append(r)
    headers3 = [c.value for c in ws3[1]]
    if "value" in headers3:
        idxv = headers3.index("value") + 1
        for cell in ws3.iter_cols(min_col=idxv, max_col=idxv, min_row=2):
            for c in cell:
                c.number_format = number_format
    if "date" in headers3:
        idxd = headers3.index("date") + 1
        for cell in ws3.iter_cols(min_col=idxd, max_col=idxd, min_row=2):
            for c in cell:
                c.number_format = date_format
    ws3.freeze_panes = "A2"
    tab3 = Table(
        displayName="RawDataTable",
        ref=f"A1:{ws3.cell(row=ws3.max_row, column=ws3.max_column).coordinate}",
    )
    tab3.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    ws3.add_table(tab3)
    # Config snapshot
    ws4 = wb.create_sheet("Konfiguration")
    cfg_text: str = ""
    try:
        import yaml

        # yaml.safe_dump returns a str; keep cfg_text annotated to satisfy mypy
        cfg_text = yaml.safe_dump(config)
    except Exception:
        cfg_text = str(config)
    cfg_lines = cfg_text.split("\n")
    for i, line in enumerate(cfg_lines, start=1):
        ws4.cell(row=i, column=1, value=line)
    # README
    ws5 = wb.create_sheet("README")
    ws5.cell(
        row=1,
        column=1,
        value="Dieses Workbook enthält das Makro-Ranking. Siehe Konfiguration für Details.",
    )
    ws5.cell(
        row=3,
        column=1,
        value="Methodik: Datenquellen, Transformationen, Standardisierung, Gewichte sind in Konfiguration beschrieben.",
    )
    # Optional: write portfolio allocations and backtest results if provided in config
    def _maybe_write_df_to_sheet(name: str, obj: Any) -> bool:
        if obj is None:
            return False
        # normalize to DataFrame
        try:
            if isinstance(obj, pd.DataFrame):
                df = obj.copy()
            elif isinstance(obj, (dict, list, pd.Series)):
                df = pd.DataFrame(obj)
            else:
                return False
        except Exception:
            return False

        wsx = wb.create_sheet(name)
        try:
            df_mapped = map_iso3_column_in_df(df.reset_index())
        except Exception:
            df_mapped = df.reset_index()
        for r in dataframe_to_rows(df_mapped, index=False, header=True):
            wsx.append(r)
        headersx = [c.value for c in wsx[1]]
        if "value" in headersx:
            idxv = headersx.index("value") + 1
            for cell in wsx.iter_cols(min_col=idxv, max_col=idxv, min_row=2):
                for c in cell:
                    c.number_format = number_format
        if "date" in headersx:
            idxd = headersx.index("date") + 1
            for cell in wsx.iter_cols(min_col=idxd, max_col=idxd, min_row=2):
                for c in cell:
                    c.number_format = date_format
        wsx.freeze_panes = "A2"
        try:
            tabx = Table(
                displayName=f"{name}Table",
                ref=f"A1:{wsx.cell(row=wsx.max_row, column=wsx.max_column).coordinate}",
            )
            tabx.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
            wsx.add_table(tabx)
        except Exception:
            # ignore table creation errors for small/empty frames
            pass
        return True

    # accept multiple possible placements in config for backward compatibility
    portfolio_obj = None
    backtest_obj = None
    # Safely extract portfolio/backtest objects from config without evaluating
    # pandas objects in boolean context (DataFrame.__bool__ raises).
    pconf = config.get("portfolio")
    if isinstance(pconf, dict):
        portfolio_obj = pconf.get("allocations") if pconf.get("allocations") is not None else pconf.get("df")
    else:
        portfolio_obj = pconf

    bconf = config.get("backtest")
    if isinstance(bconf, dict):
        backtest_obj = bconf.get("results") if bconf.get("results") is not None else bconf.get("df")
    else:
        backtest_obj = bconf

    if portfolio_obj is not None:
        try:
            # determine mapping path: prefer config.provided path then fallback to data/ file
            mapping: Dict[str, Dict[str, Any]] = {}
            try:
                mapping_path = None
                try:
                    pconf = config.get("portfolio") if isinstance(config, dict) else None
                    if isinstance(pconf, dict) and pconf.get("mapping_path"):
                        mapping_path = pconf.get("mapping_path")
                except Exception:
                    mapping_path = None
                if not mapping_path:
                    mapping_path = os.path.join(os.getcwd(), "data", "countries_iso3_map.csv")
                if os.path.exists(mapping_path):
                    import csv as _csv

                    with open(mapping_path, newline="", encoding="utf-8") as _fh:
                        rdr = _csv.DictReader(_fh)
                        for r in rdr:
                            code = (r.get("iso3") or r.get("ISO3") or r.get("code") or "").strip().upper()
                            if code:
                                mapping[code] = r
            except Exception:
                mapping = {}

            try:
                df_port: pd.DataFrame = (
                    portfolio_obj.copy() if isinstance(portfolio_obj, pd.DataFrame) else pd.DataFrame(portfolio_obj)
                )
                # ensure country column exists
                if "country" in df_port.columns:
                    df_port["_iso3"] = df_port["country"].astype(str).str.upper()
                    # add mapping columns
                    for col in ("ticker", "isin", "exchange", "currency"):
                        df_port[col] = df_port["_iso3"].apply(lambda x: mapping.get(x, {}).get(col) if mapping else None)
                    # compute delta and est_cost if prev_weight present
                    try:
                        if "prev_weight" in df_port.columns and "weight" in df_port.columns:
                            df_port["delta"] = df_port["weight"] - df_port["prev_weight"]
                            # get cost_per_unit from config.portfolio.cost_per_unit if present else default 0.001
                            cost_per_unit = 0.001
                            try:
                                pconf = config.get("portfolio") if isinstance(config, dict) else None
                                if isinstance(pconf, dict) and pconf.get("cost_per_unit") is not None:
                                    # mypy: pconf.get() may return Any | None; guard with explicit cast
                                    val = pconf.get("cost_per_unit")
                                    try:
                                        cost_per_unit = float(val) if val is not None else cost_per_unit
                                    except Exception:
                                        cost_per_unit = cost_per_unit
                            except Exception:
                                cost_per_unit = 0.001
                            df_port["est_cost"] = df_port["delta"].abs() * cost_per_unit
                    except Exception:
                        pass
                    # drop helper
                    df_port = df_port.drop(columns=["_iso3"])
                    _maybe_write_df_to_sheet("Portfolio", df_port)
                else:
                    _maybe_write_df_to_sheet("Portfolio", portfolio_obj)
            except Exception:
                _maybe_write_df_to_sheet("Portfolio", portfolio_obj)
        except Exception:
            pass
    if backtest_obj is not None:
        try:
            _maybe_write_df_to_sheet("Backtest", backtest_obj)
        except Exception:
            pass
    # Harmonize report (if provided in config or manifest)
    hrep = config.get("harmonize_report")
    if hrep is not None:
        try:
            _maybe_write_df_to_sheet("Harmonize_Report", hrep)
        except Exception:
            pass
    # Save to a temp file first, then atomically replace the target. This avoids
    # partial writes and reduces PermissionError issues when Excel has the file open.
    dirpath = os.path.dirname(path) or "."
    tmp_name = f".{os.path.basename(path)}.tmp"
    tmp_path = os.path.join(dirpath, tmp_name)
    import tempfile

    def _attempt_replace(src: str, dst: str):
        # use os.replace for atomic overwrite; let exceptions bubble
        os.replace(src, dst)

    # First try: save to tmp file in same directory (best for atomic replace)
    try:
        wb.save(tmp_path)
        try:
            _attempt_replace(tmp_path, path)
            return path
        except PermissionError:
            # couldn't replace the target (likely locked by Excel). Move the tmp
            # to a timestamped alternate filename in the same directory instead.
            import datetime

            base, ext = os.path.splitext(path)
            alt = f"{base}_{datetime.datetime.utcnow().strftime('%Y%m%dT%H%M%SZ')}{ext}"
            try:
                _attempt_replace(tmp_path, alt)
                return alt
            except Exception:
                # fallback to system temp below
                pass
    except PermissionError:
        # If we cannot write tmp in target dir, try writing to the system temp folder
        pass

    # Second try: write to system temp and attempt to replace target; if replace
    # fails, write temp file to an alternate timestamped name next to target.
    tmp2: Optional[str] = None
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_fh:
            tmp2 = tmp_fh.name
        wb.save(tmp2)
        try:
            _attempt_replace(tmp2, path)
            return path
        except PermissionError:
            import datetime

            base, ext = os.path.splitext(path)
            alt = f"{base}_{datetime.datetime.utcnow().strftime('%Y%m%dT%H%M%SZ')}{ext}"
            try:
                _attempt_replace(tmp2, alt)
                return alt
            except Exception:
                raise
    finally:
        # best-effort cleanup of tmp files
        for pth in (tmp_path, tmp2):
            try:
                if pth and os.path.exists(pth):
                    os.remove(pth)
            except Exception:
                pass
    # If we reach here, all attempts failed
    raise PermissionError(f"Failed to write Excel file to {path}")
